"""Explicit user-data loading and opt-in synthetic test-data generation."""
import csv
import json
import os
from datetime import datetime, timezone
from typing import Any, Dict, Iterable, List, Optional

from faker import Faker

from utilities.logger import get_logger

logger = get_logger("test_data")


class RequiredTestDataFieldError(ValueError):
    """Signals that a user-supplied dataset needs clarification, never fabrication."""


def _require_fields(records: Iterable[Dict[str, Any]], required_fields: Iterable[str], source: str) -> List[Dict[str, Any]]:
    validated_records = list(records)
    for row_number, record in enumerate(validated_records, start=1):
        missing = [field for field in required_fields if not str(record.get(field, "")).strip()]
        if missing:
            message = f"User-supplied data in {source}, row {row_number}, is missing required field(s): {', '.join(missing)}"
            logger.error(message)
            raise RequiredTestDataFieldError(message)
    logger.info("Validated %d user-supplied test-data record(s) from %s", len(validated_records), source)
    return validated_records


def load_user_data(file_path: str, required_fields: Optional[Iterable[str]] = None) -> List[Dict[str, Any]]:
    """Load CSV, JSON, or Excel data exactly as supplied; never synthesize missing values."""
    if not os.path.isfile(file_path):
        raise FileNotFoundError(f"User-supplied test-data file was not found: {file_path}")
    extension = os.path.splitext(file_path)[1].lower()
    logger.info("Loading user-supplied test data from %s", file_path)
    if extension == ".csv":
        with open(file_path, newline="", encoding="utf-8-sig") as data_file:
            records = list(csv.DictReader(data_file))
    elif extension == ".json":
        with open(file_path, encoding="utf-8") as data_file:
            payload = json.load(data_file)
        records = payload if isinstance(payload, list) else [payload]
        if not all(isinstance(record, dict) for record in records):
            raise ValueError("JSON test data must be an object or a list of objects")
    elif extension in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        from openpyxl import load_workbook
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            records = []
        else:
            headers = [str(value).strip() if value is not None else "" for value in rows[0]]
            records = [dict(zip(headers, row)) for row in rows[1:] if any(value is not None for value in row)]
    else:
        raise ValueError("User-supplied test data must be CSV, JSON, or Excel (.xlsx/.xlsm/.xltx/.xltm)")
    return _require_fields(records, required_fields or [], file_path)


def generate_synthetic_data(count: int = 1, locale: Optional[str] = None) -> List[Dict[str, str]]:
    """Generate unique client/engagement records only when a test explicitly opts in."""
    if count < 1:
        raise ValueError("count must be at least 1")
    faker = Faker(locale) if locale else Faker()
    run_suffix = datetime.now(timezone.utc).strftime("%Y%m%d%H%M%S%f")
    records = []
    for index in range(1, count + 1):
        suffix = f"{run_suffix}_{index:03d}"
        client_name = f"{faker.company()} {suffix}"
        records.append({"client_name": client_name, "contact_name": faker.name(),
                        "email": f"{faker.user_name()}_{suffix}@example.test", "phone": faker.msisdn(),
                        "registration_no": f"{faker.bothify(text='??######').upper()}_{suffix}",
                        "engagement_name": f"{client_name} Audit {suffix}", "unique_suffix": suffix})
    logger.info("Generated %d explicitly requested synthetic test-data record(s) | run_suffix=%s", count, run_suffix)
    return records


def get_test_data(mode: str, *, file_path: Optional[str] = None,
                  required_fields: Optional[Iterable[str]] = None,
                  count: int = 1, locale: Optional[str] = None) -> List[Dict[str, Any]]:
    """Select a mode explicitly; synthetic values are never an implicit fallback."""
    if mode == "user_supplied":
        if not file_path:
            raise ValueError("file_path is required when mode is 'user_supplied'")
        return load_user_data(file_path, required_fields)
    if mode == "synthetic":
        if file_path:
            raise ValueError("file_path is not used in explicitly selected synthetic mode")
        return generate_synthetic_data(count, locale)
    raise ValueError("mode must be explicitly set to 'user_supplied' or 'synthetic'")
